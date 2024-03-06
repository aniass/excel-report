"""Automating the Excel Report from sales data"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font 

# # Define the output path for the Excel report
OUTPUT_PATH = 'Excel_project\sales_report.xlsx'


def read_data(file):
    df = pd.read_excel(file, sheet_name='Sheet1')
    return df
 
    
def create_pivot_table(df):
    income_city = df.pivot_table(index='City',
                                values='Total',
                                columns='Product line',
                                aggfunc='sum').round(0)
    return income_city
    
    
def write_to_excel(pivot_table):
    pivot_table.to_excel(OUTPUT_PATH, sheet_name='Sales_city', startrow=4)
    
    
def format_workbook():
    wb = load_workbook(OUTPUT_PATH)
    sheet = wb.active
    sheet = wb['Sales_city']
    
     # Adjust column dimensions
    dimensions = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for col in dimensions:
        sheet.column_dimensions[col].width = 12 if col == 'A' else 20

    # Add bar chart
    chart = BarChart()
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row) 

    categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row) 

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "B11") 
    chart.title = 'Sales by city'
    chart.style = 42                
    chart.x_axis.title = "City"
    chart.y_axis.title = "Total sales"

    # Format report title and subtitle
    sheet['D1'] = 'Sales Report'
    sheet['D2'] = 'Sales by cities and products'
    for cell in ['D1', 'D2']:
        sheet[cell].font = Font('Arial', bold=True, size=24 if cell == 'D1' else 10, italic=True)
        sheet[cell].alignment = Alignment(horizontal="center")

    wb.save(OUTPUT_PATH)
    return wb


def generate_excel_report(file):
    df = read_data(file)
    pivot_table = create_pivot_table(df)
    write_to_excel(pivot_table)
    workbook = format_workbook()
    return workbook


if __name__ == '__main__':
    generate_excel_report('supermarket_sales.xlsx')
